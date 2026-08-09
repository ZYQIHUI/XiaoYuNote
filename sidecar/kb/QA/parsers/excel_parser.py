"""Excel 解析器 — 行分组合块（r1 6.2 节核心设计）"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from ...config import Config
from ...models import CellRecord, Document
from ...utils import extract_date_from_path, desensitize, is_sensitive_file

logger = logging.getLogger("kb.parsers.excel")


def parse_excel(file_path: Path, cfg: Config) -> list[Document]:
    """解析 Excel 文件，按行分组合块（r1 P0.3 升级）。

    算法：
    1. 探测表头行（非空单元格最多的行）
    2. 合并单元格前向填充
    3. 相邻数据行拼到接近 chunk.size 上限 → 一个 Document
    """
    docs: list[Document] = []
    rel_path = file_path.relative_to(cfg.root_dir).as_posix()
    date, date_source = extract_date_from_path(file_path)

    try:
        if file_path.suffix.lower() == ".xls":
            import pandas as pd
            xl = pd.ExcelFile(file_path)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Excel 解析失败: {file_path} — {e}")
        return []

    # 按 sheet 处理
    sheets: list[str] = []
    if file_path.suffix.lower() == ".xls":
        sheets = xl.sheet_names
    else:
        sheets = wb.sheetnames

    for sheet_name in sheets:
        rows_data = _read_sheet(file_path, sheet_name, cfg)
        if not rows_data:
            continue

        headers, data_rows = _detect_header(rows_data)

        # 凭据脱敏检查
        all_text = "\n".join("；".join(row) for row in data_rows)
        clean_text, desens_count = desensitize(all_text)

        sensitive = is_sensitive_file(desens_count, rel_path)
        if sensitive and cfg.security.skip_sensitive_files:
            logger.warning(f"敏感文件跳过: {rel_path}")
            continue

        # 行分组合块
        chunks = _row_group_chunk(
            headers, data_rows, rel_path, file_path,
            date, date_source, sheet_name, cfg, sensitive,
        )
        docs.extend(chunks)

    return docs


def _read_sheet(file_path: Path, sheet_name: str, cfg: Config) -> list[list[str]]:
    """读取单个 sheet 的原始文本数据。"""
    if file_path.suffix.lower() == ".xls":
        import pandas as pd
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        return [
            [str(v) if v is not None and str(v) != "nan" else "" for v in row]
            for row in df.values.tolist()
        ]

    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    result: list[list[str]] = []
    for row in ws.iter_rows(max_row=min(5000, ws.max_row), values_only=False):
        result.append([str(cell.value) if cell.value is not None else "" for cell in row])
    wb.close()
    return result


def _detect_header(rows: list[list[str]]) -> tuple[Optional[list[str]], list[list[str]]]:
    """探测表头行：取非空单元格最多的前 N 行（返回 headers + 数据行）。"""
    headers, data_rows, _ = _detect_header_with_index(rows)
    return headers, data_rows


def _detect_header_with_index(
    rows: list[list[str]],
) -> tuple[Optional[list[str]], list[list[str]], int]:
    """同 _detect_header，额外返回表头行 0-based 索引（M2 单元格坐标需要）。"""
    if not rows:
        return None, [], 0

    probe_n = min(5, len(rows))
    best_idx = 0
    best_count = -1
    for i in range(probe_n):
        non_empty = sum(1 for cell in rows[i] if cell.strip())
        if non_empty > best_count:
            best_count = non_empty
            best_idx = i

    if best_count <= 0:
        return None, rows, 0

    headers = rows[best_idx]
    data_rows = rows[best_idx + 1 :]
    return headers, data_rows, best_idx


def extract_cells(file_path: Path, cfg: Config) -> list[CellRecord]:
    """提取 Excel 非空单元格 → CellRecord 列表（M2 精确值索引）。

    与 parse_excel 同一数据源（_read_sheet + 表头探测 + 前向填充 + 脱敏），
    但保留单元格级坐标（Excel 1-based 行列 + 列字母 + 列头），
    供 /api/sheets 对单号/金额等精确值做 SQL 精确匹配。
    """
    cells: list[CellRecord] = []
    rel_path = file_path.relative_to(cfg.root_dir).as_posix()

    try:
        if file_path.suffix.lower() == ".xls":
            import pandas as pd
            xl = pd.ExcelFile(file_path)
            sheets = xl.sheet_names
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
    except Exception as e:
        logger.error(f"Excel 单元格提取失败: {file_path} — {e}")
        return []

    def col_letter(idx: int) -> str:
        """1-based 列号 → 列字母（A..Z, AA..）。"""
        letters = ""
        n = idx
        while n > 0:
            n, rem = divmod(n - 1, 26)
            letters = chr(65 + rem) + letters
        return letters

    for sheet_name in sheets:
        rows_data = _read_sheet(file_path, sheet_name, cfg)
        if not rows_data:
            continue

        headers, data_rows, header_idx = _detect_header_with_index(rows_data)

        # 敏感文件防线：与 parse_excel 一致（脱敏统计 → 敏感跳过）
        all_text = "\n".join("；".join(row) for row in data_rows)
        _, desens_count = desensitize(all_text)
        sensitive = is_sensitive_file(desens_count, rel_path)
        if sensitive and cfg.security.skip_sensitive_files:
            logger.warning(f"敏感文件跳过（cells）: {rel_path}")
            continue

        data_rows = _forward_fill(data_rows)
        header_row_excel = header_idx + 1  # 表头行 Excel 1-based 行号

        for r_i, row in enumerate(data_rows):
            excel_row = header_row_excel + 1 + r_i
            for c_i, raw in enumerate(row):
                val = raw.strip()
                if not val:
                    continue
                val_clean, _ = desensitize(val)
                hdr = headers[c_i].strip() if headers and c_i < len(headers) else ""
                cells.append(CellRecord(
                    rel_path=rel_path,
                    sheet_name=sheet_name,
                    row=excel_row,
                    col=c_i + 1,
                    col_letter=col_letter(c_i + 1),
                    header=hdr,
                    value=val_clean,
                ))

    return cells


def _forward_fill(rows: list[list[str]]) -> list[list[str]]:
    """合并单元格前向填充：空值取上方最近非空值。"""
    if not rows:
        return rows

    n_cols = max(len(r) for r in rows) if rows else 0
    for col_idx in range(n_cols):
        last_val = ""
        for row in rows:
            if col_idx < len(row):
                if row[col_idx].strip():
                    last_val = row[col_idx]
                elif last_val:
                    row[col_idx] = last_val
    return rows


def _row_group_chunk(
    headers: Optional[list[str]],
    data_rows: list[list[str]],
    rel_path: str,
    file_path: Path,
    date: int,
    date_source: str,
    sheet_name: str,
    cfg: Config,
    sensitive: bool,
) -> list[Document]:
    """相邻数据行合并为块（r1 P0.3 行分组合块）。"""
    docs: list[Document] = []
    data_rows = _forward_fill(data_rows)

    if headers is None:
        n_cols = len(data_rows[0]) if data_rows else 0
        headers = [f"col_{i}" for i in range(n_cols)]

    current_lines: list[str] = []
    start_row: Optional[int] = None
    seq = 0

    for idx, row in enumerate(data_rows):
        if all(not cell.strip() for cell in row):
            continue

        actual_row_num = idx + 2  # Excel 行号从 2 开始（1 是表头）
        line = "；".join(f"{h}={v}" for h, v in zip(headers, row) if f"{h}={v}".strip())

        if start_row is None:
            start_row = actual_row_num

        candidate = "\n".join(current_lines + [line])
        if len(candidate) > cfg.chunk.size and current_lines:
            # 当前缓冲区已满，输出为一块
            docs.append(Document(
                path=str(file_path),
                rel_path=rel_path,
                file_type="excel",
                date=date,
                date_source=date_source,
                sheet_name=sheet_name,
                row_start=start_row,
                row_end=actual_row_num - 1,
                text="\n".join(current_lines),
                sensitive=sensitive,
            ))
            current_lines = [line]
            start_row = actual_row_num
        else:
            current_lines.append(line)

    # 剩余内容
    if current_lines:
        end_row = (idx + 2) if data_rows else start_row or 2
        docs.append(Document(
            path=str(file_path),
            rel_path=rel_path,
            file_type="excel",
            date=date,
            date_source=date_source,
            sheet_name=sheet_name,
            row_start=start_row,
            row_end=end_row,
            text="\n".join(current_lines),
            sensitive=sensitive,
        ))

    return docs
