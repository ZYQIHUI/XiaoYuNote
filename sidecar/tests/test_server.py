"""sidecar 服务测试 — token 握手 / 健康检查 / 文件读写 / 索引 / SSE 问答 / 配置脱敏。

embedding 与 chat 全部 mock（不依赖外部 API）；SQLite 与文件系统为真实临时数据。
"""

import base64
import importlib
import json
from pathlib import Path

import numpy as np
import pytest
from fastapi.testclient import TestClient


class FakeEmbed:
    """批量/单条 embedding 均返回固定 1024 维归一化向量。"""

    DIM = 1024

    def __init__(self, cfg=None):
        self.rng = np.random.default_rng(7)

    def embed(self, texts):
        return [self._vec() for _ in texts]

    def embed_one(self, text):
        return self._vec()

    def _vec(self):
        v = self.rng.standard_normal(self.DIM).astype(np.float32)
        return (v / np.linalg.norm(v)).tolist()


class FakeChat:
    """chat 返回固定回答。"""

    def __init__(self, cfg=None):
        pass

    def chat(self, messages, temperature=0.3):
        return "测试回答：拼房对账清单共 2 笔单据，金额 186.50，状态已对账。"


MD_DAILY = """# 2026-7-15 日报

## 今日完成
- 拼房对账：单据 D20260721002 金额 186.50 状态 已对账
- 单据 D20260721003 金额 186.50 状态 差异
"""


@pytest.fixture
def server_env(tmp_path, monkeypatch):
    """构造临时数据目录 + 重载 server（读取新 XIAOYU_DATA_DIR），mock LLM。"""
    server, client = _make_server(tmp_path, monkeypatch)
    return server, client


@pytest.fixture
def server_env_llm(tmp_path, monkeypatch):
    """带完整 .env（key + 模型）的 server 环境：llm_ready=True（LLM 仍为 mock）。"""
    env = _kb_env_path()
    existed = env.exists()
    env.write_text(
        "OPENAI_API_KEY=sk-test-llm\n"
        "OPENAI_BASE_URL=http://127.0.0.1:9/v1\n"
        "KB_CHAT_MODEL=test-chat\n"
        "KB_EMBED_MODEL=test-embed\n",
        encoding="utf-8",
    )
    try:
        server, client = _make_server(tmp_path, monkeypatch)
        yield server, client
    finally:
        if not existed:
            env.unlink(missing_ok=True)


def _kb_env_path() -> Path:
    return Path(__file__).resolve().parent.parent / "kb" / ".env"


def _make_server(tmp_path, monkeypatch):
    """建数据目录 + 重载 server + mock LLM，返回 (server, TestClient)。"""
    # 数据目录：个人/2026-7-15/日报.md + 对账清单.xlsx
    d1 = tmp_path / "个人" / "2026-7-15"
    d1.mkdir(parents=True)
    (d1 / "日报.md").write_text(MD_DAILY, encoding="utf-8")

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "对账"
    ws.append(["单号", "金额", "状态"])
    ws.append(["D20260721002", "186.50", "已对账"])
    ws.append(["D20260721003", "186.50", "差异"])
    wb.save(str(tmp_path / "个人" / "2026-7-15" / "对账清单.xlsx"))

    monkeypatch.setenv("XIAOYU_DATA_DIR", str(tmp_path))
    import sidecar.server as server

    server = importlib.reload(server)
    monkeypatch.setattr("kb.Transformer.retriever.LLMClient", FakeEmbed)
    monkeypatch.setattr("kb.Transformer.generator.LLMClient", FakeChat)
    monkeypatch.setattr("kb.QA.embedder.LLMClient", FakeEmbed)
    return server, TestClient(server.app)


def _auth_headers(server_env) -> dict:
    server, _ = server_env
    return {"X-Token": server.SIDECAR.token}


class TestAuth:
    def test_health_with_token(self, server_env):
        _, client = server_env
        r = client.get("/api/health", headers=_auth_headers(server_env))
        assert r.status_code == 200
        body = r.json()
        assert body["status"] == "ok"
        assert body["llm_ready"] is False  # 未配置 key，降级模式
        assert body["index"]["files"] == 0

    def test_health_without_token_rejected(self, server_env):
        _, client = server_env
        assert client.get("/api/health").status_code == 401

    def test_wrong_token_rejected(self, server_env):
        _, client = server_env
        r = client.get("/api/health", headers={"X-Token": "wrong-token"})
        assert r.status_code == 401

    def test_token_is_random_and_persisted(self, server_env, tmp_path):
        server, _ = server_env
        assert len(server.SIDECAR.token) >= 32
        saved = json.loads((tmp_path / ".sidecar.json").read_text(encoding="utf-8"))
        assert saved["token"] == server.SIDECAR.token


class TestFiles:
    def test_tree_contains_both_zones(self, server_env):
        _, client = server_env
        r = client.get("/api/files/tree", headers=_auth_headers(server_env))
        assert r.status_code == 200
        names = [c["name"] for c in r.json()["children"]]
        assert "个人" in names
        assert "notes" in names

    def test_write_read_text(self, server_env):
        _, client = server_env
        h = _auth_headers(server_env)
        r = client.post("/api/files", json={"path": "notes/daily/2026-08-09.md", "content": "# 测试"}, headers=h)
        assert r.status_code == 200
        r2 = client.get("/api/files", params={"path": "notes/daily/2026-08-09.md"}, headers=h)
        assert r2.json()["content"] == "# 测试"

    def test_path_traversal_rejected(self, server_env, tmp_path):
        _, client = server_env
        h = _auth_headers(server_env)
        r = client.get("/api/files", params={"path": "../secret.txt"}, headers=h)
        assert r.status_code == 400

    def test_xlsx_roundtrip(self, server_env):
        _, client = server_env
        h = _auth_headers(server_env)
        rel = "个人/2026-7-15/对账清单.xlsx"
        r = client.get("/api/files/xlsx", params={"path": rel}, headers=h)
        assert r.status_code == 200
        b64 = r.json()["content_base64"]

        # 修改后回写
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(base64.b64decode(b64)))
        ws = wb.active
        ws.append(["D20260721004", "99.00", "已对账"])
        buf = io.BytesIO()
        wb.save(buf)
        new_b64 = base64.b64encode(buf.getvalue()).decode("ascii")

        r2 = client.put("/api/files/xlsx", json={"path": rel, "content_base64": new_b64}, headers=h)
        assert r2.status_code == 200

        r3 = client.get("/api/files/xlsx", params={"path": rel}, headers=h)
        wb2 = load_workbook(io.BytesIO(base64.b64decode(r3.json()["content_base64"])))
        assert wb2.active.max_row == 4  # 原 3 行 + 追加 1 行


class TestConfig:
    @staticmethod
    def _kb_env() -> Path:
        return _kb_env_path()

    def test_config_masks_api_key(self, server_env):
        env = self._kb_env()
        existed = env.exists()
        env.write_text("OPENAI_API_KEY=sk-test-1234567890\n", encoding="utf-8")
        try:
            _, client = server_env
            r = client.get("/api/config", headers=_auth_headers(server_env))
            assert r.status_code == 200
            masked = r.json()["api_key_masked"]
            assert masked
            assert "sk-test-1234567890" not in masked  # 不返回明文
        finally:
            if not existed:
                env.unlink(missing_ok=True)

    def test_config_write_updates_env(self, server_env):
        env = self._kb_env()
        existed = env.exists()
        try:
            _, client = server_env
            h = _auth_headers(server_env)
            r = client.post(
                "/api/config",
                json={"KB_CHAT_MODEL": "test-chat-model", "OPENAI_API_KEY": "sk-test-1234567890"},
                headers=h,
            )
            assert r.status_code == 200
            assert "OPENAI_API_KEY" in r.json()["updated"]

            # 回读：key 脱敏，模型可见
            r2 = client.get("/api/config", headers=h)
            assert r2.json()["chat_model"] == "test-chat-model"
            assert "sk-test-1234567890" not in r2.json()["api_key_masked"]
        finally:
            if not existed:
                env.unlink(missing_ok=True)


class TestIndex:
    def test_incremental_index(self, server_env):
        _, client = server_env
        h = _auth_headers(server_env)
        r = client.post("/api/index", headers=h)
        assert r.status_code == 200
        body = r.json()
        assert body["new_files"] == 2  # 日报.md + 对账清单.xlsx
        assert body["ok"] >= 1

        # 二次索引：全部跳过
        r2 = client.post("/api/index", headers=h)
        assert r2.json()["new_files"] == 0
        assert r2.json()["changed_files"] == 0

        # stats 反映入库
        r3 = client.get("/api/stats", headers=h)
        assert r3.json()["files"] == 2

    def test_index_requires_token(self, server_env):
        _, client = server_env
        assert client.post("/api/index").status_code == 401


class TestSheetsAPI:
    """表格感知问答：/api/sheets 精确值命中（M2 验收）。"""

    def test_sheets_exact_hit_after_index(self, server_env_llm):
        _, client = server_env_llm
        h = _auth_headers(server_env_llm)
        client.post("/api/index", headers=h)

        r = client.get("/api/sheets", params={"q": "D20260721002 金额"}, headers=h)
        assert r.status_code == 200
        body = r.json()
        assert body["count"] >= 1
        hits = body["hits"]
        assert any(
            "对账清单.xlsx" in hit["rel_path"]
            and hit["col_letter"] == "A"
            and hit["header"] == "单号"
            and hit["value"] == "D20260721002"
            for hit in hits
        )
        # 引用格式：文件!Sheet 单元格
        assert any("!对账 A2" in hit["ref"] for hit in hits)

    def test_sheets_money_hit(self, server_env_llm):
        _, client = server_env_llm
        h = _auth_headers(server_env_llm)
        client.post("/api/index", headers=h)
        r = client.get("/api/sheets", params={"q": "186.50"}, headers=h)
        assert r.status_code == 200
        assert any(hit["value"] == "186.50" and hit["header"] == "金额" for hit in r.json()["hits"])

    def test_sheets_empty_query_rejected(self, server_env):
        _, client = server_env
        r = client.get("/api/sheets", params={"q": "  "}, headers=_auth_headers(server_env))
        assert r.status_code == 400

    def test_sheets_without_token_rejected(self, server_env):
        _, client = server_env
        assert client.get("/api/sheets", params={"q": "x"}).status_code == 401


class TestAskSSE:
    def test_ask_stream(self, server_env_llm):
        server, client = server_env_llm
        h = _auth_headers(server_env_llm)
        # 先索引
        client.post("/api/index", headers=h)

        r = client.post("/api/ask", json={"query": "对账清单有哪些单据"}, headers=h)
        assert r.status_code == 200
        events: dict[str, list] = {}
        for line in r.text.splitlines():
            if line.startswith("event: "):
                cur = line[len("event: "):]
            elif line.startswith("data: ") and cur:
                events.setdefault(cur, []).append(json.loads(line[len("data: "):]))

        assert "error" not in events, events
        assert "done" in events, events
        refs = events["done"][0]["refs"]
        assert refs, "应返回结构化引用"
        assert any("对账清单.xlsx" in str(ref.get("source", "")) for ref in refs)
        answer = "".join(d["text"] for d in events["answer"])
        assert "拼房对账" in answer or "测试回答" in answer

    def test_ask_empty_query_rejected(self, server_env_llm):
        _, client = server_env_llm
        r = client.post("/api/ask", json={"query": "  "}, headers=_auth_headers(server_env_llm))
        assert r.status_code == 400

    def test_ask_without_token_rejected(self, server_env_llm):
        _, client = server_env_llm
        assert client.post("/api/ask", json={"query": "x"}).status_code == 401
