"""XiaoYu 侧车服务 — FastAPI（仅回环 + token 握手）。

端点：
  GET  /api/health            健康检查（含索引状态）
  POST /api/ask               SSE 流式问答（向量检索 + 生成 + 画像注入）
  POST /api/index             增量索引（rebuild=true 全量重建）
  GET  /api/stats             索引统计（文件/块/类型）
  GET  /api/config            AI 配置读取（密钥脱敏，不返回明文）
  POST /api/config            写配置（config.toml / .env 字段）
  GET  /api/files/tree        数据目录树（notes + 业务双区）
  GET  /api/files?path=       md/txt 文本读取
  POST /api/files             md/txt 文本写入
  GET  /api/files/xlsx?path=  xlsx 二进制下载
  PUT  /api/files/xlsx        xlsx 二进制上传（base64）

安全：仅绑定 127.0.0.1；所有 /api/* 端点校验 X-Token 头。
数据目录：环境变量 XIAOYU_DATA_DIR，默认 %APPDATA%\\XiaoYu（~/.xiaoyu）。
"""

from __future__ import annotations

import base64
import datetime
import json
import os
import secrets
import sys
import threading
import uuid
from pathlib import Path
from typing import Optional

# 确保 kb 子包可导入（无论从何处启动 / PyInstaller 打包后）
sys.path.insert(0, str(Path(__file__).resolve().parent))

from fastapi import FastAPI, Header, HTTPException, Request
from fastapi.responses import StreamingResponse
# 老式 .xls 转换依赖（顶层 import 确保 PyInstaller 收集）
import xlrd  # noqa: F401
import openpyxl  # noqa: F401

from kb.config import Config, load_config
from kb.QA.store import Store
from kb.QA.scanner import scan
from kb.QA.chunker import chunk_documents
from kb.QA.embedder import embed_and_store
from kb.QA.exact import search_exact
from kb.QA.parsers.excel_parser import extract_cells
from kb.Transformer.retriever import Retriever
from kb.Transformer.generator import Generator
from kb.Transformer.cli import _parse_all, _record_files
from kb.Menmory.memory import load_profile, format_profile_for_prompt

# ------------------------------------------------------------------
# 数据目录与 sidecar 配置
# ------------------------------------------------------------------

APP_NAME = "XiaoYu"
DEFAULT_PORT = 8721
BUSINESS_SOURCES = ["个人", "小组", "文档", "VDMS在线办公平台"]


def default_data_dir() -> Path:
    env = os.getenv("XIAOYU_DATA_DIR")
    if env:
        return Path(env)
    if os.name == "nt":
        base = Path(os.getenv("APPDATA", str(Path.home() / "AppData" / "Roaming")))
        return base / APP_NAME
    return Path.home() / f".{APP_NAME.lower()}"


def ensure_data_layout(data_dir: Path) -> None:
    """初始化双区数据目录：notes（XiaoYuNote 笔记体系）+ 业务区。"""
    (data_dir / "notes" / "daily").mkdir(parents=True, exist_ok=True)
    (data_dir / "notes" / "weekly").mkdir(parents=True, exist_ok=True)
    (data_dir / "notes" / "monthly").mkdir(parents=True, exist_ok=True)
    (data_dir / "notes" / "images").mkdir(parents=True, exist_ok=True)
    for src in BUSINESS_SOURCES:
        (data_dir / src).mkdir(parents=True, exist_ok=True)


class SidecarConfig:
    """sidecar 自身配置：token + 端口 + 额外知识库源（存于数据目录 .sidecar.json）。"""

    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.path = data_dir / ".sidecar.json"
        self.token: str = ""
        self.port: int = DEFAULT_PORT
        self.extra_sources: list[str] = []
        self._load_or_create()

    def _load_or_create(self) -> None:
        if self.path.exists():
            raw = json.loads(self.path.read_text(encoding="utf-8"))
            self.token = raw.get("token", "")
            self.port = int(raw.get("port", DEFAULT_PORT))
            self.extra_sources = [s for s in raw.get("extra_sources", []) if s]
        if not self.token:
            self.token = secrets.token_urlsafe(32)
            self._save()

    def set_extra_sources(self, sources: list[str]) -> None:
        """更新额外知识库文件夹（绝对路径），持久化到 .sidecar.json。"""
        self.extra_sources = [s for s in sources if s]
        self._save()

    def _save(self) -> None:
        self.path.write_text(
            json.dumps(
                {
                    "token": self.token,
                    "port": self.port,
                    "extra_sources": self.extra_sources,
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )


# ------------------------------------------------------------------
# kb 配置构造（双扫描区 + 数据目录覆盖；缺密钥时宽松降级）
# ------------------------------------------------------------------

def build_kb_config(data_dir: Path) -> Config:
    """构造 kb Config：root_dir 指向数据目录，双区 sources，db 落在数据根。

    load_config 在校验失败（缺 API key / 模型）时抛 ValueError —— 服务仍需启动，
    此时保留 .env 已注入的字段（部分配置），其余用默认值。
    """
    try:
        # 优先读数据目录 kb/.env（打包后持久化）；不存在则回退 kb 包目录（开发/测试期）
        env_path = data_dir / "kb" / ".env"
        if not env_path.exists():
            env_path = Path(__file__).resolve().parent / "kb" / ".env"
        cfg = load_config(env_path=env_path)
    except ValueError as e:
        # 未配置 API key / 模型：服务仍可启动（health/files），ask 时报明确错误
        import logging
        logging.getLogger("sidecar").warning(f"AI 配置不完整，以降级模式启动: {e}")
        cfg = Config()
        cfg.openai_api_key = os.getenv("OPENAI_API_KEY", "")
        cfg.openai_base_url = os.getenv("OPENAI_BASE_URL", "")
        cfg.chat_model = os.getenv("KB_CHAT_MODEL", "")
        cfg.embed_model = os.getenv("KB_EMBED_MODEL", "")
        cfg.embed_base_url = os.getenv("KB_EMBED_BASE_URL") or None
    cfg.root_dir = data_dir
    cfg.store.path = "kb.sqlite3"
    cfg.data.sources = BUSINESS_SOURCES + ["notes"]
    # 额外知识库文件夹（绝对路径，用户自定义）
    cfg.data.sources += SIDECAR.extra_sources
    return cfg


def llm_ready(cfg: Config) -> bool:
    return bool(cfg.openai_api_key and cfg.chat_model and cfg.embed_model)


# ------------------------------------------------------------------
# 索引
# ------------------------------------------------------------------

def run_index(data_dir: Path, rebuild: bool = False) -> dict:
    """增量（或全量重建）索引，返回统计。"""
    data_dir = Path(data_dir)
    cfg = build_kb_config(data_dir)
    store = Store(cfg)
    if rebuild:
        store.reset()
    result = scan(cfg, dry_run=False)
    docs = _parse_all(result.new_files + result.changed_files, cfg)
    _record_files(store, docs, cfg)

    # M2：重建精确值索引（cells 表）——新/变更的 Excel 文件
    for fpath in result.new_files + result.changed_files:
        if fpath.suffix.lower() in (".xlsx", ".xls"):
            cells = extract_cells(fpath, cfg)
            store.replace_cells(fpath.relative_to(data_dir).as_posix(), cells)

    chunks = chunk_documents(docs, cfg)
    stats = embed_and_store(chunks, cfg, store)  # 用真实 LLMClient（未配置时抛明确错误）
    stats["cells"] = store.cell_stats()
    stats.update(
        {
            "new_files": len(result.new_files),
            "changed_files": len(result.changed_files),
            "skipped_files": len(result.skipped_files),
            "deleted_files": len(result.deleted_rel_paths),
        }
    )
    store.close()
    return stats


# ------------------------------------------------------------------
# 文件访问（路径安全）
# ------------------------------------------------------------------

def _safe_path(data_dir: Path, rel: str) -> Path:
    """校验路径安全（防目录穿越）。

    支持两种形式：
    - 相对路径：必须位于数据目录内；
    - 绝对路径：必须位于某个用户添加的外部知识库文件夹（extra_sources）内。
    """
    p = Path(rel)
    if p.is_absolute():
        root = data_dir.resolve()
        resolved = p.resolve()
        for src in SIDECAR.extra_sources:
            src_root = Path(src).resolve()
            if resolved == src_root or src_root in resolved.parents:
                return resolved
        # 外部源可能尚未配置（如旧数据）；绝对路径不在 extra_sources 内则拒绝
        if resolved == root or root in resolved.parents:
            return resolved
        raise HTTPException(status_code=400, detail=f"非法路径: {rel}")
    p = (data_dir / rel).resolve()
    root = data_dir.resolve()
    if p != root and root not in p.parents:
        raise HTTPException(status_code=400, detail=f"非法路径: {rel}")
    return p


def _file_tree(data_dir: Path, max_depth: int = 6, root_path: Path | None = None) -> dict:
    """生成目录树（跳过隐藏文件与 .sidecar.json）。

    顶层包含数据目录本身（或指定的 root_path），外加用户添加的外部知识库文件夹
    （各自作为独立顶层项）。外部源节点携带绝对路径 `path` 字段（Flutter 端用其
    读写文件）；数据目录内节点不带 path，由 Flutter 端按相对路径拼接。
    """
    base = root_path or data_dir
    # 指定根目录且是外部源 → 子节点用绝对 path；数据目录内子目录 → rel_root 前缀
    is_external_root = root_path is not None and root_path != data_dir and root_path.resolve() not in [p for p in data_dir.resolve().parents] and root_path.resolve() != data_dir.resolve()
    tree: dict = {"name": base.name, "type": "dir", "children": []}
    if root_path is not None and not is_external_root and root_path != data_dir:
        # 指定了根目录（相对数据目录内的子目录）：记录其相对路径供前端拼接
        try:
            tree["rel_root"] = base.relative_to(data_dir).as_posix()
        except ValueError:
            pass
    skip = {".sidecar.json", "kb"}

    def walk(node: dict, path: Path, depth: int, abs_base: Path | None = None) -> None:
        if depth > max_depth:
            return
        for child in sorted(path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
            if child.name.startswith(".") or child.name in skip:
                continue
            if child.is_dir():
                sub: dict = {"name": child.name, "type": "dir", "children": []}
                if abs_base is not None:
                    sub["path"] = str(child.resolve())
                walk(sub, child, depth + 1, abs_base)
                node["children"].append(sub)
            elif child.suffix.lower() in (".md", ".txt", ".xlsx", ".xls", ".docx", ".pdf"):
                entry: dict = {
                    "name": child.name,
                    "type": "file",
                    "size": child.stat().st_size,
                }
                if abs_base is not None:
                    entry["path"] = str(child.resolve())
                node["children"].append(entry)

    walk(tree, base, 0, abs_base=base if is_external_root else None)
    if root_path is not None:
        # 指定根目录时只返回该目录的树，不附加外部源
        return tree

    # 外部知识库文件夹：作为独立顶层项（name=目录名，path=绝对路径）
    for src in SIDECAR.extra_sources:
        sp = Path(src)
        if not sp.is_dir():
            continue
        ext_node: dict = {"name": sp.name, "type": "dir", "path": str(sp.resolve()), "children": []}
        walk(ext_node, sp, 0, abs_base=sp.resolve())
        tree["children"].append(ext_node)

    return tree

# ------------------------------------------------------------------
# FastAPI 应用
# ------------------------------------------------------------------

app = FastAPI(title="XiaoYu sidecar", version="0.1.0")
INDEX_LOCK = threading.Lock()

# 应用级数据目录（首次 import 时确定，测试可覆盖）
DATA_DIR = Path(os.getenv("XIAOYU_DATA_DIR") or default_data_dir())
ensure_data_layout(DATA_DIR)
SIDECAR = SidecarConfig(DATA_DIR)


def _resolve_univer_dist() -> Optional[Path]:
    """定位 univer_app/dist（WebView 表格前端，ESM 需经 http 提供）。

    候选路径（按存在性优先）：
      1. 打包：exe 同目录 univer_app/dist（本文件位于 exe 同目录的 sidecar/ 下）
      2. 开发：仓库 spring_note/univer_app/dist
    """
    here = Path(__file__).resolve()
    # 打包 onedir：.../Release/sidecar/_internal/sidecar/server.py → 上溯 4 级到 Release
    for up in range(0, 6):
        base = here
        for _ in range(up):
            base = base.parent
        cand = base / "univer_app" / "dist"
        if (cand / "index.html").is_file():
            return cand
    # 开发：仓库根 spring_note/univer_app/dist
    cand = here.parent.parent / "spring_note" / "univer_app" / "dist"
    if (cand / "index.html").is_file():
        return cand
    return None


_UNIVER_DIST = _resolve_univer_dist()
if _UNIVER_DIST is not None:
    from fastapi.staticfiles import StaticFiles

    app.mount(
        "/univer",
        StaticFiles(directory=str(_UNIVER_DIST), html=True),
        name="univer",
    )


def _auth(x_token: Optional[str] = Header(default=None)) -> None:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")


@app.get("/api/health")
def health(x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    cfg = build_kb_config(DATA_DIR)
    store = Store(cfg)
    stats = store.stats()
    store.close()
    return {
        "status": "ok",
        "data_dir": str(DATA_DIR),
        "llm_ready": llm_ready(cfg),
        "index": stats,
    }


@app.post("/api/ask")
async def ask(request: Request, x_token: Optional[str] = Header(default=None)) -> StreamingResponse:
    """SSE 流式问答：retrieved → answer（分块）→ done。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")

    body = await request.json()
    query = str(body.get("query", "")).strip()
    if not query:
        raise HTTPException(status_code=400, detail="query 不能为空")
    k = body.get("k")
    path_prefix = body.get("path")
    date_range = body.get("date_range")  # [start_YYYYMMDD, end_YYYYMMDD]

    cfg = build_kb_config(DATA_DIR)
    if not llm_ready(cfg):
        return _sse_error("AI 未配置：请先设置 OPENAI_API_KEY / 模型（.env 或 /api/config）")

    def event_stream():
        try:
            retriever = Retriever(cfg)
            generator = Generator(cfg)
            profile = load_profile()
            profile_text = format_profile_for_prompt(profile)

            yield _sse("status", {"phase": "retrieving"})
            chunks = retriever.topk(
                query,
                k=k,
                path_prefix=path_prefix,
                date_range=tuple(date_range) if date_range else None,
            )
            if not chunks:
                yield _sse("retrieved", {"chunks": []})
                yield _sse("answer", {"text": "知识库中未找到相关内容"})
                yield _sse("done", {"refs": []})
                return

            yield _sse(
                "retrieved",
                {"chunks": [{"source": c.get("rel_path"), "score": round(float(c.get("score", 0)), 4)} for c in chunks]},
            )

            today_str = datetime.date.today().isoformat()
            answer, refs = generator.generate(query, chunks, profile_text=profile_text, today_str=today_str)

            # 分块推送（按段切分，模拟流式）
            seg_size = 120
            answer_clean = (answer or "").strip()
            for i in range(0, len(answer_clean), seg_size):
                yield _sse("answer", {"text": answer_clean[i : i + seg_size]})
            yield _sse("done", {"refs": refs})
        except Exception as e:
            yield _sse("error", {"message": f"{type(e).__name__}: {e}"})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/index")
def index(rebuild: bool = False, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    with INDEX_LOCK:
        return run_index(DATA_DIR, rebuild=rebuild)


@app.get("/api/stats")
def stats(x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    cfg = build_kb_config(DATA_DIR)
    store = Store(cfg)
    s = store.stats()
    meta_model = store.get_meta("embedding_model")
    s["embedding_model"] = meta_model
    s["cells"] = store.cell_stats()
    store.close()
    return s


@app.get("/api/sheets")
def sheets(q: str, path: Optional[str] = None, top_n: int = 10,
           x_token: Optional[str] = Header(default=None)) -> dict:
    """表格感知问答：从问题提取精确值 token → cells 精确匹配 + 表头上下文。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    if not q.strip():
        raise HTTPException(status_code=400, detail="q 不能为空")
    cfg = build_kb_config(DATA_DIR)
    store = Store(cfg)
    try:
        hits = search_exact(store, q, path_prefix=path, top_n=top_n)
    finally:
        store.close()
    return {"query": q, "hits": hits, "count": len(hits)}


# ------------------------------------------------------------------
# /api/config：读写 kb 配置（config.toml + .env），密钥脱敏
# ------------------------------------------------------------------

def _mask(key: str) -> str:
    if len(key) <= 8:
        return "*" * len(key)
    return key[:4] + "*" * (len(key) - 8) + key[-4:]


@app.get("/api/config")
def get_config(x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    cfg = build_kb_config(DATA_DIR)
    return {
        "data_dir": str(DATA_DIR),
        "sources": cfg.data.sources,
        "extra_sources": SIDECAR.extra_sources,
        "openai_base_url": cfg.openai_base_url,
        "api_key_masked": _mask(cfg.openai_api_key) if cfg.openai_api_key else "",
        "chat_model": cfg.chat_model,
        "embed_model": cfg.embed_model,
        "embed_base_url": cfg.embed_base_url,
        "chunk_size": cfg.chunk.size,
        "chunk_overlap_ratio": cfg.chunk.overlap_ratio,
        "top_k": cfg.retrieval.top_k,
        "max_context_chars": cfg.retrieval.max_context_chars,
    }


@app.post("/api/config")
async def set_config(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    # 持久化到数据目录（开发期与打包后一致，重启不丢失）
    kb_dir = DATA_DIR / "kb"
    kb_dir.mkdir(parents=True, exist_ok=True)
    env_path = kb_dir / ".env"

    env_lines = env_path.read_text(encoding="utf-8").splitlines() if env_path.exists() else []
    env_map: dict[str, str] = {}
    for line in env_lines:
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            env_map[k.strip()] = v.strip().strip("'\"")
    for k in ("OPENAI_API_KEY", "OPENAI_BASE_URL", "KB_CHAT_MODEL", "KB_EMBED_MODEL", "KB_EMBED_BASE_URL"):
        if k in body and body[k] is not None:
            env_map[k] = str(body[k])
    # 额外知识库文件夹（绝对路径）——持久化到 .sidecar.json
    if "extra_sources" in body:
        raw = body["extra_sources"]
        if isinstance(raw, list):
            SIDECAR.set_extra_sources([str(s) for s in raw])
    with open(env_path, "w", encoding="utf-8") as f:
        f.write("# XiaoYu .env（由 sidecar 管理）\n")
        for k, v in env_map.items():
            f.write(f"{k}={v}\n")
    return {"ok": True, "updated": sorted(env_map.keys())}


# ------------------------------------------------------------------
# /api/files：目录树 + md/txt 文本 + xlsx 二进制
# ------------------------------------------------------------------

@app.get("/api/files/tree")
def files_tree(x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    return _file_tree(DATA_DIR)


@app.get("/api/dirs")
def list_dirs(x_token: Optional[str] = Header(default=None)) -> dict:
    """可浏览的根目录列表：数据目录笔记区 + 业务区 + 外部知识库文件夹。

    每个目录带 path（相对数据目录或绝对路径）、label（显示名）。
    """
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    dirs: list[dict] = []

    # 数据目录内：notes 三区 + 业务文件夹（存在才列）
    for rel in ["notes/daily", "notes/weekly", "notes/monthly"] + BUSINESS_SOURCES:
        p = DATA_DIR / rel
        if p.is_dir():
            dirs.append({
                "path": rel,
                "label": f"{DATA_DIR.name}/{rel}",
                "root": rel,
                "abs_path": str(p),
            })
    # 数据目录本身（含 notes 根）
    dirs.insert(0, {"path": "", "label": DATA_DIR.name, "root": "", "abs_path": str(DATA_DIR)})
    # 外部知识库文件夹（绝对路径）
    for src in SIDECAR.extra_sources:
        sp = Path(src)
        if sp.is_dir():
            dirs.append({
                "path": str(sp),
                "label": str(sp),
                "root": str(sp),
                "abs_path": str(sp),
            })
    return {"dirs": dirs}


@app.get("/api/files/tree/root")
def files_tree_root(root: str, x_token: Optional[str] = Header(default=None)) -> dict:
    """以指定根目录生成文件树（root=相对数据目录路径或外部源绝对路径）。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    base = _safe_path(DATA_DIR, root) if root else DATA_DIR
    if not base.is_dir():
        raise HTTPException(status_code=404, detail=f"目录不存在: {root}")
    return _file_tree(DATA_DIR, root_path=base)


@app.get("/api/files")
def read_text(path: str, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    p = _safe_path(DATA_DIR, path)
    if not p.is_file() or p.suffix.lower() not in (".md", ".txt"):
        raise HTTPException(status_code=404, detail="仅支持 md/txt 文件")
    return {"path": path, "content": p.read_text(encoding="utf-8", errors="replace")}


@app.post("/api/files")
async def write_text(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    path = str(body.get("path", ""))
    content = str(body.get("content", ""))
    p = _safe_path(DATA_DIR, path)
    if p.suffix.lower() not in (".md", ".txt"):
        raise HTTPException(status_code=400, detail="仅支持 md/txt 写入")
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content, encoding="utf-8")
    return {"ok": True, "path": path, "size": len(content.encode("utf-8"))}


@app.post("/api/files/create")
async def create_file(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    """新建文件（md/txt 可带内容；xlsx 可选 base64）。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    path = str(body.get("path", ""))
    p = _safe_path(DATA_DIR, path)
    if p.exists():
        raise HTTPException(status_code=400, detail=f"已存在: {path}")
    ext = p.suffix.lower()
    if ext not in (".md", ".txt", ".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="仅支持 md/txt/xlsx 新建")
    p.parent.mkdir(parents=True, exist_ok=True)
    if ext in (".xlsx", ".xls"):
        b64 = str(body.get("content_base64", ""))
        if b64:
            try:
                data = base64.b64decode(b64, validate=True)
            except Exception:
                raise HTTPException(status_code=400, detail="content_base64 非法")
        else:
            # 最小空 xlsx（openpyxl 兼容）
            import io
            from openpyxl import Workbook
            wb = Workbook()
            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
        p.write_bytes(data)
    else:
        content = str(body.get("content", ""))
        p.write_text(content, encoding="utf-8")
    return {"ok": True, "path": path}


@app.post("/api/files/mkdir")
async def create_dir(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    """新建文件夹（相对数据目录或外部源内）。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    path = str(body.get("path", ""))
    p = _safe_path(DATA_DIR, path)
    if p.exists():
        raise HTTPException(status_code=400, detail=f"已存在: {path}")
    p.mkdir(parents=True, exist_ok=True)
    return {"ok": True, "path": path}


@app.post("/api/files/delete")
async def delete_file(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    """删除文件或文件夹（文件夹递归删除）。"""
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    path = str(body.get("path", ""))
    p = _safe_path(DATA_DIR, path)
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"不存在: {path}")
    if p.is_dir():
        import shutil
        shutil.rmtree(p)
    else:
        p.unlink()
    return {"ok": True, "path": path}


def _xls_to_xlsx_bytes(raw: bytes) -> bytes:
    """老式 .xls（BIFF）→ .xlsx 字节（xlrd 读 + openpyxl 写）。"""
    import io as _io

    import xlrd
    import openpyxl

    wb = xlrd.open_workbook(file_contents=raw)
    out = openpyxl.Workbook()
    out.remove(out.active)
    for ws in wb.sheets():
        nws = out.create_sheet(title=ws.name[:31] or "Sheet")
        # 列宽（xlrd colinfo_map: {colx: Colinfo(width)}）
        try:
            for colx, info in ws.colinfo_map.items():
                if info and getattr(info, "width", None):
                    # xlrd 宽度单位是 1/256 字符宽；openpyxl 用字符数
                    nws.column_dimensions[chr(65 + colx) if colx < 26 else openpyxl.utils.get_column_letter(colx + 1)].width = max(info.width / 256.0, 8)
        except Exception:
            pass
        # 行高
        try:
            for rowx, info in ws.rowinfo_map.items():
                if info and getattr(info, "height", None):
                    nws.row_dimensions[rowx + 1].height = max(info.height / 20.0, 15)
        except Exception:
            pass
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                v = cell.value
                if v is None:
                    continue
                # 日期单元格：xlrd 返回 float（Excel 序列号），带 ctype=3 标记
                if cell.ctype == 3 and isinstance(v, float):
                    from datetime import datetime, timedelta
                    dt = datetime(1899, 12, 30) + timedelta(days=v)
                    nws.cell(row=r + 1, column=c + 1).value = dt
                    nws.cell(row=r + 1, column=c + 1).number_format = "yyyy/m/d"
                else:
                    nws.cell(row=r + 1, column=c + 1).value = v
    buf = _io.BytesIO()
    out.save(buf)
    return buf.getvalue()


@app.get("/api/files/xlsx")
def read_xlsx(path: str, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    p = _safe_path(DATA_DIR, path)
    if not p.is_file() or p.suffix.lower() not in (".xlsx", ".xls"):
        raise HTTPException(status_code=404, detail="仅支持 xlsx/xls 文件")
    raw = p.read_bytes()
    # 老式 .xls（BIFF）转 .xlsx，前端 exceljs 才能读取
    if p.suffix.lower() == ".xls" and not raw.startswith(b"PK"):
        try:
            raw = _xls_to_xlsx_bytes(raw)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"xls 转换失败: {e}")
    data = base64.b64encode(raw).decode("ascii")
    return {"path": path, "content_base64": data, "size": p.stat().st_size}


@app.put("/api/files/xlsx")
async def write_xlsx(request: Request, x_token: Optional[str] = Header(default=None)) -> dict:
    if x_token != SIDECAR.token:
        raise HTTPException(status_code=401, detail="无效或缺失的 token")
    body = await request.json()
    path = str(body.get("path", ""))
    b64 = str(body.get("content_base64", ""))
    p = _safe_path(DATA_DIR, path)
    if p.suffix.lower() not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="仅支持 xlsx/xls 写入")
    p.parent.mkdir(parents=True, exist_ok=True)
    try:
        data = base64.b64decode(b64, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="content_base64 非法")
    p.write_bytes(data)
    return {"ok": True, "path": path, "size": len(data)}


# ------------------------------------------------------------------
# SSE 工具
# ------------------------------------------------------------------

def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def _sse_error(message: str) -> StreamingResponse:
    import json as _json
    body = f"event: error\ndata: {_json.dumps({'message': message}, ensure_ascii=False)}\n\n"
    return StreamingResponse(iter([body]), media_type="text/event-stream")


def _find_free_port(start: int, tries: int = 20) -> int:
    """从 start 起探测可用端口（先绑定后释放，避免 uvicorn 直接退出）。"""
    import socket

    for port in range(start, start + tries):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                continue
    raise RuntimeError(f"无可用的回环端口（{start}~{start + tries - 1} 均被占用）")


def main() -> None:
    import uvicorn

    port = _find_free_port(SIDECAR.port)
    if port != SIDECAR.port:
        print(f"端口 {SIDECAR.port} 被占用，改用 {port}")
        # 写回实际端口，Flutter 客户端从 .sidecar.json 读取
        SIDECAR.port = port
        SIDECAR._save()
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="info")


if __name__ == "__main__":
    main()
