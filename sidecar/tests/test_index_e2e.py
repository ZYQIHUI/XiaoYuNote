"""index 全链路端到端测试 — 扫描→解析→记录→切块→向量化→入库（embedding 用 mock）

覆盖：文件记录落库（mtime_ns 回归）、增量跳过、文件变更后重新入库无残留。
"""

import numpy as np
import pytest
from pathlib import Path
from types import SimpleNamespace

from kb.config import Config
from kb.QA.store import Store
from kb.QA.scanner import scan
from kb.QA.chunker import chunk_documents
from kb.QA.embedder import embed_and_store
from kb.Transformer.cli import _parse_all, _record_files


class FakeLLM:
    """模拟 embedding：固定 1024 维 L2 归一化向量（与配置 dim 一致）。"""

    def __init__(self, dim: int = 1024):
        self.dim = dim
        self.rng = np.random.default_rng(42)

    def embed(self, texts: list[str]) -> list[list[float]]:
        return [self._vec() for _ in texts]

    def _vec(self) -> list[float]:
        v = self.rng.standard_normal(self.dim).astype(np.float32)
        return (v / np.linalg.norm(v)).tolist()


MD_DAILY = """# 2026-7-15 日报

## 今日完成
- 完成费控云「产品预订」模块功能测试，覆盖正向与异常场景，共输出 12 条用例
- 发现缺陷 #1023：预订数量输入 0 时页面无提示，已提单跟踪

## 问题记录
- 缺陷 #1023 已关联需求文档，待开发修复后执行回归验证
"""


@pytest.fixture
def project(tmp_path):
    """构造临时项目：个人/ 数据源 + md + xlsx，返回 (cfg, store, root)。"""
    from openpyxl import Workbook

    d1 = tmp_path / "个人" / "2026-7-15"
    d1.mkdir(parents=True)
    (d1 / "日报.md").write_text(MD_DAILY, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "拼房明细"
    ws.append(["房间号", "入住人", "日期", "房费", "结算状态"])
    ws.append(["A301", "张三", "2026-07-20", "600.00", "已结算"])
    ws.append(["A301", "李四", "2026-07-20", "600.00", "未结算"])
    d2 = tmp_path / "个人" / "2026-7-23"
    d2.mkdir(parents=True)
    wb.save(str(d2 / "拼房大表.xlsx"))

    cfg = Config(root_dir=tmp_path)
    cfg.data.sources = ["个人"]
    store = Store(cfg)
    return SimpleNamespace(cfg=cfg, store=store, root=tmp_path)


class TestIndexE2E:
    """端到端：全量入库 → 增量跳过 → 文件变更重入库。"""

    def _full_index(self, project) -> tuple[list, dict]:
        """执行一次完整入库，返回 (chunks, stats)。"""
        result = scan(project.cfg, dry_run=False)
        docs = _parse_all(result.new_files + result.changed_files, project.cfg)
        _record_files(project.store, docs, project.cfg)
        chunks = chunk_documents(docs, project.cfg)
        stats = embed_and_store(chunks, project.cfg, project.store, llm=FakeLLM())
        return chunks, stats

    def test_full_index_flow(self, project):
        """全链路：文件记录、向量入库、检索查询、模型登记。"""
        chunks, stats = self._full_index(project)

        assert len(chunks) > 0
        assert stats["ok"] == len(chunks)
        assert stats["failed"] == 0

        # 文件记录落库（回归：cli.py 曾用不存在的 st_ns，此处必须是 st_mtime_ns）
        records = project.store.get_all_files()
        assert len(records) == 2
        for rec in records.values():
            assert rec.mtime_ns > 0, "mtime_ns 未正确落库"
            assert rec.size_bytes > 0
            assert len(rec.file_hash) == 64

        # 检索查询：chunk 带 embedding 与 rel_path
        rows = project.store.get_all_chunks_for_retrieval()
        assert len(rows) == len(chunks)
        for r in rows:
            assert r["embedding"] is not None
            assert r["rel_path"].startswith("个人/")

        # 换模型防护：首次登记模型，换模型后应拦截
        assert project.store.check_embedding_model("BAAI/bge-m3") is True
        assert project.store.check_embedding_model("other/model") is False

    def test_second_scan_incremental_skip(self, project):
        """增量：首次入库后再次扫描，全部跳过，块数不变。"""
        self._full_index(project)
        chunk_count_1 = len(project.store.get_all_chunks_for_retrieval())

        result = scan(project.cfg, dry_run=False)
        assert len(result.new_files) == 0
        assert len(result.skipped_files) == 2

        # 再跑一次入库（无新内容）不应产生新块
        docs = _parse_all(result.new_files + result.changed_files, project.cfg)
        chunks = chunk_documents(docs, project.cfg)
        embed_and_store(chunks, project.cfg, project.store, llm=FakeLLM())
        chunk_count_2 = len(project.store.get_all_chunks_for_retrieval())
        assert chunk_count_2 == chunk_count_1

    def test_file_change_reindex_no_stale(self, project):
        """文件变更：重入库后旧 chunk 不残留（块数应与重新解析一致）。"""
        self._full_index(project)

        # 修改日报内容
        md_path = project.root / "个人" / "2026-7-15" / "日报.md"
        md_path.write_text(
            "# 2026-7-15 日报\n\n## 今日完成\n- 内容已更新：新增预算模块测试 5 条用例并记录缺陷 #1050\n",
            encoding="utf-8",
        )

        # 变更文件重新入库
        result = scan(project.cfg, dry_run=False)
        assert len(result.changed_files) == 1
        docs = _parse_all(result.changed_files, project.cfg)
        _record_files(project.store, docs, project.cfg)
        chunks = chunk_documents(docs, project.cfg)
        embed_and_store(chunks, project.cfg, project.store, llm=FakeLLM())

        rows = project.store.get_all_chunks_for_retrieval()
        total = len(rows)
        # 检索到的块必须全部来自当前磁盘内容：缺陷 #1050 存在，缺陷 #1023 已不在
        texts = " ".join(r["text"] for r in rows)
        assert "#1050" in texts
        assert "#1023" not in texts
        # 块总数不得超过当前解析结果（新 md 1 块 + xlsx 1 块）
        assert total <= 2

    def test_xlsx_change_no_stale(self, project):
        """拼房大表（后续文件）变更：旧 chunk 不得残留（seq 偏移回归）。"""
        from openpyxl import load_workbook

        self._full_index(project)

        # 修改 xlsx：追加一行拼房结算
        xlsx_path = project.root / "个人" / "2026-7-23" / "拼房大表.xlsx"
        wb = load_workbook(str(xlsx_path))
        ws = wb.active
        ws.append(["B205", "王五", "2026-07-23", "450.00", "已结算"])
        wb.save(str(xlsx_path))

        result = scan(project.cfg, dry_run=False)
        assert len(result.changed_files) == 1
        docs = _parse_all(result.changed_files, project.cfg)
        _record_files(project.store, docs, project.cfg)
        chunks = chunk_documents(docs, project.cfg)
        embed_and_store(chunks, project.cfg, project.store, llm=FakeLLM())

        rows = project.store.get_all_chunks_for_retrieval()
        texts = " ".join(r["text"] for r in rows)
        # 新行在库中，且无重复的旧 xlsx 块（md 1 + xlsx 1 = 2）
        assert "B205" in texts
        assert len(rows) == 2
